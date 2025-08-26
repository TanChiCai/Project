import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime

book = load_workbook('course.xlsx')  # conncect to excel
sheet = book.active


class GPA_Calculator:
    is_running=False

    # create main window
    def _init_(self):
        if GPA_Calculator.is_running:
            return
        GPA_Calculator.is_running=True

        self.window = tk.Tk()  # creates the main window of the application using the Tk class from the tkinter library
        self.window.title("GPA Calculator")  # main window name is GPA Calculator

        self.container = tk.Frame(
            self.window)  # create a frame for hold up button,label which is attached to the main window self.window.
        self.container.grid(row=0, column=0,
                            sticky="n")  # place the frame into window by using grid() in row 0 column 0 and top(north)
        self.window.grid_rowconfigure(0, weight=1)  # resizable for row based on size of window,1 times of resizable
        self.window.grid_columnconfigure(0,
                                         weight=1)  # resizable for column based on size of window,1 times of resizable
        self.container.grid_rowconfigure(0, weight=1)  # resizable for row based on size of window,1 times of resizable
        self.container.grid_columnconfigure(0,
                                            weight=1)  # resizable for column based on size of window,1 times of resizable
        self.window.tkraise()  # bring the window in front
        self.window.attributes("-topmost", 1)  # always stay on top of all other windows  1 is True
        self.window.iconbitmap('C:/Users/user/OneDrive/Desktop/python-Assignment/GPA.ico')
        self.current_page = "main_page"  # Track the current page

        # Create the first page (initial frame)
        self.create_main_page()  # call the main page function

        # Bind the Enter key to trigger the on_first_run method
        self.window.bind('<Return>', lambda event: self.on_first_run())  # make sure once press enter then can run
        self.i = False  # as a flag  to control on_first_run function

        # Start the main loop
        self.window.mainloop()

    # test for run begining
    def on_first_run(self):

        self.focused_widget = self.window.focus_get()  # Get the widget that has focus

        if not self.i:  # make sure only run one times only for beginning
            # Check if focused widget and all entries in the current row are empty
            if self.focused_widget:  # make sure only press enter key directly without doing any process will prompt error
                self.create_main_page()  # call the main page back
                messagebox.showerror("Error", "Invalid input.Please reenter again")  # show error for without any input

                self.window.bind('<Return>',
                                 lambda event: self.on_first_run())  # make sure this function  will run again if press

    # create main page(GPA Calculator Input Interface)
    def create_main_page(self):  # create main page

        # self has an attribute called display_result_page.   exists and is not None or invalid).
        if hasattr(self, 'display_result_page') and self.display_result_page:  # if have,then hide the page
            self.display_result_page.grid_forget()

        if hasattr(self,
                   'calculation_page') and self.calculation_page.winfo_exists():  ##if only have,then hide the page
            self.calculation_page.grid_forget()


        self.current_page = "main_page"  # Track the current page
        self.window = tk.Tk()  # creates the main window of the application using the Tk class from the tkinter library
        # Create a frame for the main page
        self.main_page = tk.Frame(self.window, width=800, height=600,
                                  bg="lightgray")  # Set a background color for visibility

        self.main_page.grid(row=0, column=0, sticky="n", padx=10, pady=10)  # Add 10 pixels of padding

        # Configure grid behavior
        self.window.grid_rowconfigure(0, weight=1)  # Allow the first row to expand
        self.window.grid_columnconfigure(0, weight=1)  # Allow the first column to expand

        tk.Label(self.main_page, text="WELCOME TO TARC DCS GPA CALCULATOR", bg="dark blue", fg="white",
                 font=("Arial", 25)).grid(
            row=0, column=0, columnspan=3, pady=10, sticky="nsew")  # Switch to the results page

        tk.Label(self.main_page, text="GPA Calculator Input Interface", bg="blue", fg="white", font=("Arial", 16)).grid(
            row=1, column=0, columnspan=3, pady=10, sticky="nsew")  # Switch to the results page

        # Table setup
        self.rows = 1

        columns = 3
        headers = ["Course", "Coursework Mark", "Final Exam Mark"]

        # Create table headers
        for col in range(columns):
            header_label = tk.Label(self.main_page, text=headers[col], relief="solid", bg="lightblue",
                                    font=("Arial", 16))
            header_label.grid(row=2, column=col, padx=5, pady=5, sticky="ew")

        # Create initial table cells as Entry widgets
        self.table_entries = []
        self.create_table_row()

        # Buttons for main page
        self.Back_to_main_menu = tk.Button(self.main_page, text="Back", command=self.Back_to_main_menu,font=("Arial", 16))
        self.Back_to_main_menu.grid(row=self.rows + 4, column=0, columnspan=2, pady=10)

        self.calculate_button = tk.Button(self.main_page, text="Calculate", command=self.calculatebutton,
                                          font=("Arial", 16))
        self.calculate_button.grid(row=self.rows + 4, column=0, columnspan=7, pady=10)

        self.clear_button = tk.Button(self.main_page, text="Clear", command=self.clearbutton, font=("Arial", 16))
        self.clear_button.grid(row=self.rows + 4, column=1, columnspan=2, pady=10)

        self.add_button = tk.Button(self.main_page, text="+ Add More Course", command=self.add_coursebutton,
                                    font=("Arial", 16))
        self.add_button.grid(row=self.rows + 3, column=0, columnspan=5, pady=10)

        self.main_page.bind('<Return>', lambda event: self.Enter_key())

    # back to main menu
    def Back_to_main_menu(self):
        GPA_Calculator.is_running=False
        import MAIN_MENU_TO_DO_LIST_MANAGER_KAITERNG  # Import the main menu to show it again
        self.window.withdraw()  # Hide the GPA Calculator window

        MAIN_MENU_TO_DO_LIST_MANAGER_KAITERNG.open_main_menu()  # Call the main menu function to display it

    # create row for the table in GPA Calculator Input Interface
    def create_table_row(self):

        columns = 3
        row_entries = []
        self.Course = []

        self.Course.append(" ")  # Add the cell value to the choices list,for cancel the coursecode choice

        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):  # Iterate through rows
            if row[0] and row[1]:  # Ensure both columns have values
                cell = f"{row[0]} {row[1]}"  # Combine both columns

                self.Course.append(cell)  # Append the combined value to the list

        for j in range(columns):
            entry_var = tk.StringVar()

            if j == 0:

                # Configure the font style for the combobox entry and dropdown list
                self.dropdown = ttk.Combobox(self.main_page, textvariable=entry_var, values=self.Course, width=30,
                                             font=("Arial", 16))
                self.dropdown.grid(row=self.rows + 2, column=j, padx=5, pady=5, sticky="ew")
                self.dropdown.config(state='readonly')  # Make the dropdown read-only
                self.dropdown.grid(row=self.rows + 2, column=j, padx=5, pady=5)
                self.dropdown.bind("<Right>", self.move_to_next_entry_and_between_number)
                self.dropdown.bind("<Up>", self.move_to_previous_row)
                self.dropdown.bind("<Key>", self.on_press_key)  # Bind the key press event

                row_entries.append(self.dropdown)

            else:
                # Create an Entry widget and bind the Enter key event
                self.entry = tk.Entry(self.main_page, textvariable=entry_var, width=30, justify="center",
                                      relief="solid",
                                      font=("Arial", 16))
                self.entry.grid(row=self.rows + 2, column=j, padx=5, pady=5)
                self.entry.bind("<Right>", self.move_to_next_entry_and_between_number)
                self.entry.bind("<Left>", self.move_to_previous_entry_and_between_number)
                self.entry.bind("<Down>", self.move_to_next_row)
                self.entry.bind("<Up>", self.move_to_previous_row)
                self.entry.bind("<Key>", self.on_press_key)  # Bind the key press event

                row_entries.append(self.entry)

            if j == columns - 1:  # Ensure the row is added after the last column is processed
                self.table_entries.append(row_entries)

        # get the position of widget(cursor)

    def get_widget_position(self, widget):
        # Find the row and column of the given widget
        for r, row in enumerate(self.table_entries):
            for c, entry in enumerate(row):
                if entry == widget:
                    return r, c
        return None, None  # Return None if widget not found

    # Move right (within the same row)
    def move_to_next_entry_and_between_number(self, event):
        current_widget = event.widget

        if isinstance(current_widget, tk.Entry):  # widget used for text input
            current_text = current_widget.get()  # what inside the widget
            current_index = current_widget.index(tk.INSERT)  # Get current cursor position

            # Find the index of the next space (word boundary) from the current position
            next_space = current_text.find(' ', current_index)  # ' ' index

            # if no found,the next space is -1
            if next_space == -1:
                # If no space is found, move the cursor to the end of the text
                current_widget.icursor(current_index)  # set the cursor (insertion point) at a specific position

            if current_index == len(current_text):  # make in sure end of number
                row, col = self.get_widget_position(current_widget)

                if row is None or col is None:
                    return  # The widget is not found in the table

                if col < len(self.table_entries[row]) - 1:
                    next_widget = self.table_entries[row][col + 1]
                    next_widget.focus_set()

            return  # Exit the function after moving the cursor within an entry

    # Move left (within the same row)
    def move_to_previous_entry_and_between_number(self, event):
        current_widget = event.widget

        if isinstance(current_widget, tk.Entry):
            current_index = current_widget.index(tk.INSERT)  # Get the current cursor position
            if current_index > 0:  # Check if the cursor isn't at the beginning
                current_widget.icursor(current_index)  # Move the cursor to the left
            if current_index == 0:  # cursor at the begining
                row, col = self.get_widget_position(current_widget)

                if row is None or col is None:
                    return  # The widget is not found in the table

                if col > 0:
                    previous_widget = self.table_entries[row][col - 1]
                    previous_widget.focus_set()
                else:
                    return  # No movement if at the edge

    # Move down (same column, next row)
    def move_to_next_row(self, event):
        current_widget = event.widget
        row, col = self.get_widget_position(current_widget)

        if current_widget in self.table_entries[-1]:
            if event.keysym == "Down":  # Check if Down Arrow key is pressed
                self.add_coursebutton()

        if row < len(self.table_entries) - 1:
            next_widget = self.table_entries[row + 1][col]
            next_widget.focus_set()
        else:
            return  # No movement if at the bottom edge

    # Move up (same column, previous row)
    def move_to_previous_row(self, event):
        current_widget = event.widget
        row, col = self.get_widget_position(current_widget)

        if row is None or col is None:
            return  # The widget is not found in the table

        if row > 0:
            previous_widget = self.table_entries[row - 1][col]
            previous_widget.focus_set()
        else:
            return  # No movement if at the top edge

    # for the action of press enter key when put the cursor inside entry
    def on_press_key(self, event):
        self.window.unbind('<Return>')  # This will stop the empty method from being called when Enter is pressed

        self.entry.bind("<Key>", self.on_press_key)  # Bind the key press event
        self.dropdown.bind("<Key>", self.on_press_key)  # Bind the key press event

        if event.keysym == "Return":
            self.calculatebutton()

            return  # Exit the function after calling calculatebutton

    # for the action of directly press enter key without put cursor inside,not only that,make sure after go back still can run by pressing enter key directly
    def Enter_key(self):

        focused_widget = self.window.focus_get()
        if isinstance(focused_widget, tk.Entry) or isinstance(focused_widget, ttk.Combobox):
            # Do nothing if the focused widget is Entry or Dropdown
            return
        if self.current_page == "main_page":
            self.calculatebutton()
            return
        elif self.current_page == "Calculate":
            self.display_results()
            return
        elif self.current_page == "displayresult":
            return

    # Add the row of table so that allow multiple course enter
    def add_coursebutton(self):

        self.rows += 1

        if self.rows > len(self.Course) - 1:
            messagebox.showinfo("Info", "You have reached the maximum limit of course!")
            return
        self.create_table_row()

        self.Back_to_main_menu.grid_forget()
        self.Back_to_main_menu.grid(row=self.rows + 4, column=0, columnspan=2, pady=10)

        # Update button positions
        self.calculate_button.grid_forget()
        self.calculate_button.grid(row=self.rows + 4, column=0, columnspan=7, pady=10)

        self.clear_button.grid_forget()
        self.clear_button.grid(row=self.rows + 4, column=1, columnspan=2, pady=10)

        self.add_button.grid_forget()
        self.add_button.grid(row=self.rows + 3, column=0, columnspan=5, pady=10)

        self.window.update_idletasks()  # the chnage will see quickly

    dialog_open=False
    # clear for the all row and data if press the clear button
    def clearbutton(self):

        # Confirm action with the user
        if not self.dialog_open:
            self.dialog_open=True
            message = messagebox.askquestion('Info', 'Do you really want to clear',parent=self.window)
            self.dialog_open=False
            if message == "yes":
                for row in self.table_entries:  # Iterate through each row in the table
                    for widget in row:  # Iterate through each widget (Entry or Combobox) in the row
                        if isinstance(widget, tk.Entry):  # Check if it's an Entry widget
                            widget.delete(0, 'end')  # Clear the content of the Entry
                        if isinstance(widget, ttk.Combobox):  # Check if it's a Combobox widget
                            widget.set('')  # Clear the Combobox selection

                while len(self.table_entries) > 1:
                    row_to_delete = self.table_entries.pop()  # Remove the last row from the list
                    for widget in row_to_delete:
                        widget.destroy()  # Destroy the widgets in that row

                    # Reset the row counter to 1
                self.rows = 1
    # check whether couseworkmark and Finalexammark is number or not
    def is_number(self, couseworkmark, Finalexammark):
        if couseworkmark == '' or Finalexammark == '':
            return False
        try:

            Couseworkmark = float(couseworkmark)
            Finalexammarks = float(Finalexammark)

            return True
        except:
            return False

    # page to show the Academic Grade Tracker,which is the process of calculation gpa
    def calculatebutton(self):

        if hasattr(self, 'display_result_page') and self.display_result_page:
            self.display_result_page.grid_forget()

        self.current_page = "Calculate"
        self.calculation_page = tk.Frame(self.window)
        self.calculation_page.grid(row=0, column=0, sticky="nsew")
        self.calculation_page.focus_set()

        course = []
        couseworkmarks = []
        Finalexammarks = []
        Totalcredithours = []
        grades = []
        Grade_points = []
        descriptions = []
        self.data_has_entries = False
        credit_hour_get = 0

        for row in self.table_entries:

            self.i = True  # once become True front will no run
            Courses = row[0].get()  # Get the value of each entry in the row

            couseworkmark = row[1].get()  # Get the value of each entry in the row
            Finalexammark = row[2].get()  # Get the value of each entry in the row

            if Courses and couseworkmark and Finalexammark:
                self.data_has_entries = True  # Mark that we found a row with data

            if (Courses == " " or not Courses) and not couseworkmark and not Finalexammark:
                if not self.data_has_entries and self.i:
                    if self.rows == 1:
                        self.showerror()  # Show an error message
                        continue
                    else:
                        self.data_has_entries = True
                        continue

            if self.data_has_entries == True:
                if (Courses == " " or not Courses) and not couseworkmark and not Finalexammark:
                    continue

            if not self.checkerror(Courses, couseworkmark, Finalexammark):
                self.showerror()
                return  # Stop processing if any row data is invalid

            if (float(couseworkmark) < 50 or float(couseworkmark) > 100) or (
                    float(Finalexammark) < 0 or float(Finalexammark) > 100):
                messagebox.showerror("Error",
                                     "Invalid input. Please make sure coursework mark >=50 and <=100,final exam mark>=0 and <=100.")
                self.calculation_page.destroy()
                self.main_page.tkraise()
                return

            Finalmarks = round(float(couseworkmark) / 2 + float(Finalexammark) / 2, 2)  # ensure two decimal point

            # Append the data to respective lists
            course.append(Courses)
            couseworkmarks.append(couseworkmark)
            Finalexammarks.append(Finalexammark)

            for i in range(len(course)):
                for j in range(i + 1, len(course)):
                    if course[i] == course[j]:
                        self.showerror()
                        return

            if Finalmarks >= 90 and Finalmarks <= 100:
                grade = "A+"
                Grade_point = 4.00
                description = "High Distinction"

            elif Finalmarks >= 80 and Finalmarks <= 89:
                grade = "A"
                Grade_point = 4.00
                description = "Distinction"

            elif Finalmarks >= 75 and Finalmarks <= 79:
                grade = "A-"
                Grade_point = 3.67
                description = "Distinction"

            elif Finalmarks >= 70 and Finalmarks <= 74:
                grade = "B+"
                Grade_point = 3.33
                description = "Merit"

            elif Finalmarks >= 65 and Finalmarks <= 69:
                grade = "B"
                Grade_point = 3.00
                description = "Merit"

            elif Finalmarks >= 60 and Finalmarks <= 64:
                grade = "B-"
                Grade_point = 2.67
                description = "Merit"

            elif Finalmarks >= 55 and Finalmarks <= 59:
                grade = "C+"
                Grade_point = 2.33
                description = "Pass"

            elif Finalmarks >= 50 and Finalmarks <= 54:
                grade = "C"
                Grade_point = 2.00
                description = "Pass"

            elif Finalmarks >= 0 and Finalmarks <= 49:
                grade = "F"
                Grade_point = 0.00
                description = "Fail"

            if float(Finalexammark) < 40:
                grade = "F"
                Grade_point = 0.00
                description = "Fail"
            grades.append(grade)
            Grade_points.append(Grade_point)
            descriptions.append(description)
            n = 0
            courses = []

            for row in sheet.iter_rows(values_only=True):
                if row[0] and row[1]:  # Ensure both columns have values
                    cell = f"{row[0]} {row[1]}"  # Combine both columns

                    if Courses == cell:

                        courses.append(Courses)
                        credithour = float(Courses[7])
                        Totalcredithours.append(credithour)

                        if grade == "F":
                            credithour = 0
                        credit_hour_get += credithour
                        break

        tk.Label(self.calculation_page, text="Academic Grade Tracker", bg="blue", fg="white", font=("Arial", 16)).grid(
            row=0, column=0, columnspan=3, pady=10, sticky="nsew")  # Switch to the results page

        totalgradepoint = []
        totalcredit = 0
        overallgpa = 0

        style = ttk.Style()
        style.configure("Treeview", font=("Arial", 8))  # Change font size here
        style.configure("Treeview.Heading", font=("Arial", 13))  # Change header font size

        # Create Treeview widget for displaying course data
        columns = ["Course", "Credit", "Grade", "Grade Points", "Total Grade Point", "Description"]

        self.tree = ttk.Treeview(self.calculation_page, columns=columns, show="headings", height=8)
        self.tree.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=250)

        # self.tree.column("#0", width=500)  # Example of fixed column width

        v_scrollbar = ttk.Scrollbar(self.calculation_page, orient="vertical", command=self.tree.yview)
        v_scrollbar.grid(row=1, column=1, sticky="ns")  # Place next to Treeview on the right
        self.tree.configure(yscrollcommand=v_scrollbar.set)

        # Add horizontal scrollbar
        h_scrollbar = ttk.Scrollbar(self.calculation_page, orient="horizontal", command=self.tree.xview)
        h_scrollbar.grid(row=2, column=0, sticky="ew")  # Ensure it is below the Treeview
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        # Configure grid weights for dynamic resizing
        self.calculation_page.grid_rowconfigure(1, weight=0)  # Allow Treeview to expand vertically
        self.calculation_page.grid_rowconfigure(2, weight=0)  # Horizontal scrollbar row
        self.calculation_page.grid_columnconfigure(0, weight=1)  # Main content column (Treeview)
        self.calculation_page.grid_columnconfigure(0, weight=1)  # Allow Treeview to expand horizontally
        total_grade_point = 0

        for i in range(len(Totalcredithours)):  # CANCEL for " "

            totalgradepoint.append(Totalcredithours[i] * Grade_points[i])
            totalcredit += Totalcredithours[i]
        for i in range(len(totalgradepoint)):
            total_grade_point += totalgradepoint[i]
        overallgpa = (total_grade_point) / totalcredit

        style = ttk.Style()
        style.configure("Treeview", highlightthickness=0, bd=0, font=('Arial', 10))

        # Configure the tag for "F" grade rows: light pink background and red foreground
        self.tree.tag_configure("light red",
                                foreground="red")  # Light pink background, red text for "F" grades

        for i in range(len(course)):
            if grades[i] == "F":
                self.tree.insert("", "end", values=(
                    course[i], Totalcredithours[i], grades[i], Grade_points[i], totalgradepoint[i],
                    descriptions[i]), tags=("light red",))

            else:
                self.tree.insert("", "end", values=(
                    course[i], Totalcredithours[i], grades[i], Grade_points[i], totalgradepoint[i],
                    descriptions[i]))

        # Create a frame for additional widgets below the treeview
        frame = tk.Frame(self.calculation_page)
        frame.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=10, pady=5)

        # Add labels or other widgets inside the frame
        credit_label = tk.Label(frame, text="Total Credits: {}".format(totalcredit), font=("Arial", 13),
                                anchor="w")  # Set font family and size (e.g., 20 points))
        credit_label.grid(row=0, column=0, columnspan=2, sticky="w", pady=5)

        gpa_label = tk.Label(frame, text="Overall GPA: {:.4f}".format(overallgpa), font=("Arial", 13), anchor="w")
        gpa_label.grid(row=1, column=0, columnspan=2, sticky="w", pady=5)

        self.results_data = (overallgpa, credit_hour_get, course, grades)

        # Create "Display" button and set up its grid placement
        self.displayresults = tk.Button(
            frame, text="Display",
            command=self.display_results,
            font=("Arial", 16)
        )
        self.displayresults.grid(row=0, column=0, columnspan=4, pady=5, padx=825)  # Place below Treeview
        self.Back = tk.Button(
            frame, text="Back",
            command=lambda: self.back("main"),
            font=("Arial", 16)
        )

        self.Back.grid(row=0, column=0, columnspan=2, pady=5, padx=625)  # Place below the Display button

        if hasattr(self, 'main_page') and self.main_page:
            self.main_page.grid_forget()
        self.window.bind('<Return>', lambda event: self.Enter_key())

    # make sure  course_code, courseworkmark, finalexammark is correct
    def checkerror(self, course_code, courseworkmark, finalexammark):
        self.i = True

        try:
            # Check for empty marks specifically, in case this isn't handled by is_number

            if course_code == " " and courseworkmark and finalexammark:
                return False

            if '.' in courseworkmark or '.' in finalexammark:
                courseworkmark_integer_part, courseworkmark_decimal_part = courseworkmark.split('.')
                finalexammark_integer_part, finalexammark_decimal_part = finalexammark.split('.')

                if len(courseworkmark_decimal_part) > 2 or len(finalexammark_decimal_part) > 2:
                    return False

            if not courseworkmark or not finalexammark or not course_code:
                return False  # Indicate invalid data due to empty marks

            # Check if the marks are valid numbers
            if not self.is_number(courseworkmark, finalexammark):
                return False  # Indicate invalid data due to invalid marks

            return True  # Return True if all checks pass

        except:
            return False  # Indicate invalid data due to an unexpected error

    # show error for the any unexpected input
    def showerror(self):

        messagebox.showerror("Error", "Invalid input.Please reenter again")
        self.calculation_page.destroy()
        self.main_page.tkraise()
        return

    # display the result of calculation process,like total credit point earned,gpa and cgpa
    def display_results(self, event=None):
        current_semester = datetime.today().strftime("%Y%m")  # Year and Month format

        current_date = datetime.today().strftime("%d %B %Y")
        # Hide all pages first
        self.hidepage()

        self.current_page = "displayresult"
        if hasattr(self, 'results_data'):  # retrieve the result data
            overallgpa, credit_hour_get, course, grades = self.results_data

        self.display_result_page = tk.Frame(self.window)
        self.display_result_page.grid(row=0, column=0, sticky="nsew")

        # Set up the scrollbar
        scrollbar = ttk.Scrollbar(self.display_result_page, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")  # Place scrollbar on the right side and fill vertically
        self.tree.configure(yscrollcommand=scrollbar.set)

        h_scrollbar = ttk.Scrollbar(self.display_result_page, orient="horizontal", command=self.tree.xview)
        h_scrollbar.pack(side="bottom", fill="x")  # Add scrollbar to bottom
        self.tree.configure(xscrollcommand=h_scrollbar.set)

        # Header Section
        header_frame = tk.Frame(self.display_result_page, bg="blue")
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="Examination Result", bg="blue", fg="white", font=("Arial", 16)).pack(pady=10)

        semester = "Semester: " + current_semester
        terminatingsession = "Terminating Session: " + str(int(current_semester) + 400)

        # Semester and Overall Section
        tk.Label(self.display_result_page, text=semester, font=("Arial", 12)).pack()
        tk.Label(self.display_result_page, text=current_date, font=("Arial", 10)).pack()

        # GPA and CGPA Section
        gpa_frame = tk.Frame(self.display_result_page)
        gpa_frame.pack(pady=10)
        tk.Label(gpa_frame, text="GPA:{:.4f}".format(overallgpa), font=("Arial", 12)).grid(row=0, column=0, padx=20)
        tk.Label(gpa_frame, text="CGPA:{:.4f}".format(overallgpa), font=("Arial", 12)).grid(row=0, column=1, padx=20)
        tk.Label(gpa_frame, text="Credits Earned:{:.1f}".format(credit_hour_get), font=("Arial", 12)).grid(row=1,
                                                                                                           column=0,
                                                                                                           padx=20)
        tk.Label(gpa_frame, text="Total Credits Earned:{:.1f}".format(credit_hour_get), font=("Arial", 12)).grid(row=1,
                                                                                                                 column=1,
                                                                                                                 padx=20)

        # Result Table Section
        columns = ("Course", "Grade")
        self.tree = ttk.Treeview(self.display_result_page, columns=columns, show="headings", height=8)
        self.tree.pack(pady=10)

        for col in columns:
            self.tree.heading(col, text=col, anchor="center")
            self.tree.column(col, width=500, anchor="center")

        # Create and configure style for Treeview (apply general style)
        style = ttk.Style()
        style.configure("Treeview", highlightthickness=0, bd=0, font=('Arial', 10))

        # Configure the tag for "F" grade rows: light pink background and red foreground
        self.tree.tag_configure("light red",
                                foreground="red")  # Light pink background, red text for "F" grades

        # Loop to insert course and grade rows
        for i in range(len(course)):
            if grades[i] == "F":
                # Insert the "F" grade row with the "light red" tag, which will make the text red and the background light pink
                self.tree.insert("", "end", values=(course[i], grades[i]), tags=("light red",))
            else:
                # Insert rows without any specific tag for non-"F" grades
                self.tree.insert("", "end", values=(course[i], grades[i]))

        # Termination Session
        tk.Label(self.display_result_page, text=terminatingsession, font=("Arial", 12), fg="red").pack(
            pady=10)

        self.Back = tk.Button(self.display_result_page, text="Back", command=lambda: self.back("Calculation"),
                              font=("Arial", 16))

        self.Back.pack(pady=5, padx=650, side='top', anchor='center')

    # go back to previous page when press
    def back(self, page_name):

        # Hide all pages first
        self.hidepage()

        # Show the desired page
        if page_name == "Calculation":
            self.calculation_page.grid(row=0, column=0, sticky="nsew")
            self.current_page = "Calculate"  # Update the current page

        if page_name == "main":
            self.main_page.grid(row=0, column=0, sticky="n", padx=10, pady=10)
            self.current_page = "main_page"  # Update the current page

    # hide the unnecessary page to prevent overlapping of page
    def hidepage(self):
        # hide the page that is present
        if hasattr(self, 'display_result_page') and self.display_result_page:
            self.display_result_page.grid_forget()

        if hasattr(self, 'calculation_page') and self.calculation_page:
            self.calculation_page.grid_forget()

        if hasattr(self, 'main_page') and self.main_page:
            self.main_page.grid_forget()

        self.current_page = None


GPA_Calculator()